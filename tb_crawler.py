import time
import json
import random
import os
import platform
import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from urllib.parse import quote
from pyquery import PyQuery as pq
from selenium.webdriver import ActionChains
import lxml
from fake_useragent import UserAgent
from openpyxl import load_workbook


class TaobaoSpider:
    def __init__(self):
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("--disable-blink-features=AutomationControlled")
        user_ag = UserAgent().random
        self.options.add_argument('user-agent=%s'%user_ag)
        self.get_user_info()
        if self.is_windows():
            self.options.binary_location = self.chromepath

        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option("useAutomationExtension", False)
        self.options.add_experimental_option("prefs", {"prfile.managed_default_content_setting.images": 2})
        self.browser = webdriver.Chrome(options=self.options)

        # 反爬设置  webdriver为 undefined
        self.browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
        			Object.defineProperty(navigator, 'webdriver', {
        			  get: () => undefined
        			})
        		  """
        })
        self.wait = WebDriverWait(self.browser, 10)
        self.loginurl = "https://login.taobao.com/member/login.jhtml"
        self.trytime = 0

    # 从config.json 文件中读取相关的配置
    def get_user_info(self):
        with open("config.json", "r", encoding="utf-8") as f:
            config = json.load(f)

        self.username = config["username"]
        self.password = config["password"]
        self.category = config["category"]
        self.keyword = config["keyword"]
        self.maxpage = config["maxpage"]
        self.chromepath = config["chromepath"]
        
        # print(self.password)
    
    # 滑块滑动
    def slide_block(self):
        time.sleep(4)
        try:
            slide = self.browser.find_element_by_id("nc_1_n1z")
            print("正在破解滑块")
        except Exception as error_msg:
            time.sleep(5)
            print(error_msg)
            return
        else:
            action = ActionChains(self.browser)
            action.click_and_hold(slide)
            sum = 0
            while True:
                x = random.randint(20,40)
                action.move_by_offset(x,0)
                time.sleep((random.randint(1,2))/10)
                sum += x
                if sum >= 260:
                    break
            action.release().perform()
            time.sleep(2)
            parser = lxml.html.etree.HTML(self.browser.page_source)
            element = parser.xpath("//div[@id='J_ItemList']/div[@class='product  ']")
            if len(element) == 0:
                flash_button = self.browser.find_element_by_xpath("//span[@class='nc-lang-cnt']/a")
                time.sleep(3)
                flash_button.click()
                if self.count_slide == 2:
                    input("滑块等待命令指示...")
                    print("破解滑块成功!")
                    return
                self.count_slide += 1
                self.slide_block()
            else:
                print("破解滑块成功!")
                return

    # 判断是否是windows系统
    def is_windows(self):
        return platform.system().lower() == "windows"

    # 获取不同平台的chromedriver的路径
    def get_chromedriver_exe_path(self):
        ret = "./bin/mac/chromedriver"
        if self.is_windows():
            ret = "./bin/win/chromedriver"
        return ret

    def login(self):
        self.browser.get(self.loginurl)
        try:
            # 找到用户名输入框,输入账号密码并登录
            username_input = self.wait.until(EC.presence_of_element_located((By.ID, "fm-login-id")))
            username_input.send_keys(self.username)

            password_input = self.wait.until(EC.presence_of_element_located((By.ID, "fm-login-password")))
            password_input.send_keys(self.password)

            login_button = self.wait.until(EC.presence_of_element_located((By.CLASS_NAME, "fm-button")))
            login_button.click()

            # ”site-nav-login-info-nick” 找到名字标签并打印内容
            taobao_name_tag = self.wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, "site-nav-login-info-nick ")))
            print(f"登录成功:{taobao_name_tag.text}")
            time.sleep(random.randint(10,20))

        except Exception as e:
            print(e)
            self.browser.close()
            print("登录失败")

    # 爬取相关的内容
    def crawl(self, cate="", key=""):
        if(cate==""):
            cate = self.category
        if(key==""):
            key=self.keyword
        self.category = cate
        self.keyword = key
        for i in range(1, self.maxpage + 1):
            self.index_page(i)

    # 爬取一页的内容
    def index_page(self, index):
        try:
            url = "https://s.taobao.com/search?q=" + quote(self.keyword) + "&s=" + str((index-1)*44)
            self.browser.get(url)
            self.count_slide = 0
            time.sleep(5)

            if index == 1:
                print('创建sheet')
                self.excelfile = "%s.xlsx" % self.category
                try:
                    df = pd.DataFrame({}, index=[1], columns=['image', 'price', 'deal', 'title', 'shop', 'location'])
                    writer = pd.ExcelWriter('%s.xlsx' % self.category)
                    book = load_workbook('%s.xlsx' % self.category)
                    writer.book = book
                    df.to_excel(writer, sheet_name='%s' % self.keyword, index=False)
                    writer.save()
                    writer.close()
                except:
                    df = pd.DataFrame({}, index=[1], columns=['image', 'price', 'deal', 'title', 'shop', 'location'])
                    df.to_excel(self.excelfile, sheet_name='%s' % self.keyword, index=False)
                print('创建成功')
            print("正在爬取第", index, "页")

            self.wait.until(
                EC.text_to_be_present_in_element((By.CSS_SELECTOR, "#mainsrp-pager li.item.active span"), str(index)))
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#mainsrp-itemlist div.item"), ))
            self.slide_block()
            self.wait.until(
                EC.text_to_be_present_in_element((By.CSS_SELECTOR, "#mainsrp-pager li.item.active span"), str(index)))
            self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#mainsrp-itemlist div.item"), ))
            self.get_product()
            self.trytime = 0

        except TimeoutException:
            self.trytime += 1
            if self.trytime >= 5:
                if (index < self.maxpage):
                    self.index_page(index + 1)
                    self.trytime = 0

    def get_product(self):
        html = self.browser.page_source
        doc = pq(html)
        items = doc("#mainsrp-itemlist .items .item").items()
        # print("-----------"*40 + "\n")
        # print(items)
        # print("-----------" * 40 + "\n")

        item_list = [item for item in items]

        dfs = []

        for index in range(len(item_list)):
            item = item_list[index]
            product = {
                'image': item.find('.pic .img').attr('data-src'),
                'price': item.find('.price strong').text(),
                'deal': item.find('.deal-cnt').text(),
                'title': item.find('.title').text(),
                'shop': item.find('.shop span').text(),
                'location': item.find('.location').text()
            }
            # print(product)
            df = pd.DataFrame(product, index=[1], columns=['image', 'price', 'deal', 'title', 'shop', 'location'])
            dfs.append(df)
        self.save2excel(dfs)

    # 把每页的数据保存到excel中
    def save2excel(self, dfs):
        print('正在将结果保存到excel中')
        total_df = pd.concat(dfs, ignore_index=True)
        # print(total_df)
        if os.path.exists(self.excelfile):
            before_df = pd.read_excel(self.excelfile, sheet_name='%s' % self.keyword)
            total_df = pd.concat([before_df, total_df], ignore_index=True).drop_duplicates()
        
        workbook = load_workbook(self.excelfile)
        if workbook.get_sheet_names()[0] == self.keyword:
            workbook.close()
            print('仅有一张sheet')
            total_df.to_excel(self.excelfile, sheet_name='%s' % self.keyword, index=False)
        else:
            workbook.remove_sheet(workbook[self.keyword])
            workbook.save(self.excelfile)
            print(workbook.get_sheet_names())
            with pd.ExcelWriter(self.excelfile, mode='a') as writer:  
                total_df.to_excel(writer, sheet_name='%s' % self.keyword, index=False)
        print("保存成功")

if __name__ == "__main__":
    spider = TaobaoSpider()
    spider.login()
    '''
    k = {
        '糖类':['口香糖', '棒棒糖', '硬糖', '软糖', '奶糖', '压片糖果'],
        '膨化食品':['甜甜圈', '爆米花', '咪咪'],
        '果干类':['果干', '坚果', '瓜子', '话梅', '肉铺'],
        '冲泡类':['泡腾片', '速溶奶茶', '速溶咖啡', '泡面', '麦片'],
        '能量补充类':['小面包', '饼干', '膨化饼干', '旺仔小馒头'],
        '即食类':['即食鸡', '辣条', '即食藕片海带结']
    }   
    for cate in k:
        for key in k[cate]:
            print('正在保存' + cate + '的' + key)
            spider.crawl(cate,key)
            time.sleep(20)
    '''
    spider.crawl()
